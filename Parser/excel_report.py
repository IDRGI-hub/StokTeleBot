import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from config import ARTICLES


# Импортируем ARTICLES из config.py
def generate_excel_report(data_dir="Parser/data", base_filename="output", output_file="weekly_report.xlsx"):
    files = sorted([
        f for f in os.listdir(data_dir)
        if f.startswith(base_filename) and f.endswith(".json")
    ])

    date_labels = []
    stock_data = {}

    # Собираем данные по датам
    for file in files:
        date_str = file[len(base_filename) + 1:-5]
        date_labels.append(date_str)

        with open(os.path.join(data_dir, file), "r", encoding="utf-8") as f:
            json_data = json.load(f)

        for article_key, item in json_data.items():
            if article_key not in stock_data:
                stock_data[article_key] = {}

            if isinstance(item, dict):
                total = sum(item.get("details", {}).values())
                stock_data[article_key][date_str] = total
            else:
                stock_data[article_key][date_str] = item

    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    header_font = Font(bold=True)
    category_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Заголовок
    headers = ["Название"] + date_labels
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = header_font
        ws.cell(row=1, column=col_num).alignment = center_alignment

    row_idx = 2

    # Группировка по категориям
    categories = {}
    for article, meta in ARTICLES.items():
        cat = meta.get("category", "Прочее")
        categories.setdefault(cat, []).append(article)

    for cat_name in sorted(categories.keys()):
        # Вставка категории как разделителя
        ws.append([cat_name])
        ws.cell(row=row_idx, column=1).font = category_font
        row_idx += 1

        for article in categories[cat_name]:
            product_info = ARTICLES.get(article, {})
            name = product_info.get("name", article)

            row = [name]
            for date in date_labels:
                qty = stock_data.get(str(article), {}).get(date, 0)
                row.append(qty)

            ws.append(row)
            row_idx += 1

        ws.append([])  # пустая строка между категориями
        row_idx += 1

    # Сохраняем файл
    output_path = os.path.join(data_dir, output_file)
    wb.save(output_path)
    return output_path
